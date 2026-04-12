from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.core.mail import send_mail
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from .models import Produit, Mouvement


# ─── Alertes ──────────────────────────────────────────────────────────────────

def envoyer_alerte_stock(produit):
    if produit.quantite < 5:
        sujet = f"⚠️ Alerte stock faible : {produit.nom}"
        message = f"""Bonjour,

Le produit suivant a un stock critique :

  Produit  : {produit.nom}
  Stock    : {produit.quantite} unité(s) restante(s)
  Seuil    : 5 unités

Veuillez réapprovisionner ce produit dès que possible.

— StockPro""".strip()
        try:
            send_mail(
                sujet,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [settings.ALERTE_EMAIL_DESTINATAIRE],
                fail_silently=False,
            )
        except Exception as e:
            print(f"Erreur envoi email alerte : {e}")


# ─── Auth ─────────────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, "Identifiants incorrects. Veuillez réessayer.")
    return render(request, 'gestion_stock/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required(login_url='/login/')
def home_view(request):
    if request.user.is_superuser:
        return redirect('admin_dashboard')
    groups = request.user.groups.values_list('name', flat=True)
    if 'gestionnaire' in groups:
        return redirect('gestionnaire')
    elif 'secretaire' in groups:
        return redirect('secretaire')
    return redirect('admin_dashboard')


# ─── Dashboard ────────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def admin_dashboard(request):
    produits = Produit.objects.all()
    total_produits = produits.count()
    valeur_stock = sum(p.prix * p.quantite for p in produits)
    alertes = produits.filter(quantite__lt=5).count()
    derniers_mouvements = Mouvement.objects.select_related('produit', 'utilisateur').order_by('-date')[:10]
    for p in produits:
        p.valeur_totale = p.prix * p.quantite
    return render(request, 'gestion_stock/admin_dashboard.html', {
        'produits': produits,
        'total_produits': total_produits,
        'valeur_stock': valeur_stock,
        'alertes': alertes,
        'derniers_mouvements': derniers_mouvements,
    })


# ─── Produits ─────────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def gestionnaire_view(request):
    produits = Produit.objects.all().order_by('nom')
    return render(request, 'gestion_stock/gestionnaire.html', {'produits': produits})


@login_required(login_url='/login/')
def ajouter_produit(request):
    if request.method == 'POST':
        nom = request.POST.get('nom', '').strip()
        quantite = int(request.POST.get('quantite', 0))
        prix = request.POST.get('prix', 0)
        if nom:
            produit = Produit.objects.create(nom=nom, quantite=quantite, prix=prix)
            if quantite > 0:
                Mouvement.objects.create(
                    produit=produit, type_mouvement='entree',
                    quantite=quantite, utilisateur=request.user,
                    note='Stock initial à la création'
                )
            envoyer_alerte_stock(produit)
            messages.success(request, f'Produit "{nom}" ajouté avec succès.')
        else:
            messages.error(request, 'Le nom du produit est requis.')
    return redirect('gestionnaire')


@login_required(login_url='/login/')
def modifier_produit(request, pk):
    produit = get_object_or_404(Produit, pk=pk)
    if request.method == 'POST':
        ancienne_qte = produit.quantite
        produit.nom = request.POST.get('nom', produit.nom).strip()
        nouvelle_qte = int(request.POST.get('quantite', produit.quantite))
        produit.prix = request.POST.get('prix', produit.prix)
        produit.quantite = nouvelle_qte
        produit.save()
        diff = nouvelle_qte - ancienne_qte
        if diff > 0:
            Mouvement.objects.create(
                produit=produit, type_mouvement='entree',
                quantite=diff, utilisateur=request.user, note='Ajustement de stock'
            )
        elif diff < 0:
            Mouvement.objects.create(
                produit=produit, type_mouvement='sortie',
                quantite=abs(diff), utilisateur=request.user, note='Ajustement de stock'
            )
        envoyer_alerte_stock(produit)
        messages.success(request, f'Produit "{produit.nom}" mis à jour.')
    return redirect('gestionnaire')


@login_required(login_url='/login/')
def supprimer_produit(request, pk):
    produit = get_object_or_404(Produit, pk=pk)
    if request.method == 'POST':
        nom = produit.nom
        produit.delete()
        messages.success(request, f'Produit "{nom}" supprimé.')
    return redirect('gestionnaire')


# ─── Mouvements ───────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def mouvement_view(request):
    if request.method == 'POST':
        produit_id = request.POST.get('produit')
        type_mvt = request.POST.get('type_mouvement')
        quantite = int(request.POST.get('quantite', 0))
        note = request.POST.get('note', '')
        produit = get_object_or_404(Produit, pk=produit_id)
        if type_mvt == 'sortie' and quantite > produit.quantite:
            messages.error(request, f'Stock insuffisant. Disponible : {produit.quantite}')
        elif quantite <= 0:
            messages.error(request, 'La quantité doit être supérieure à 0.')
        else:
            if type_mvt == 'entree':
                produit.quantite += quantite
            else:
                produit.quantite -= quantite
            produit.save()
            Mouvement.objects.create(
                produit=produit, type_mouvement=type_mvt,
                quantite=quantite, utilisateur=request.user, note=note
            )
            envoyer_alerte_stock(produit)
            messages.success(request, f'Mouvement enregistré pour "{produit.nom}".')
        return redirect('mouvements')

    produits = Produit.objects.all().order_by('nom')
    mouvements = Mouvement.objects.select_related('produit', 'utilisateur').order_by('-date')
    return render(request, 'gestion_stock/mouvements.html', {
        'produits': produits,
        'mouvements': mouvements,
    })


# ─── Secrétaire ───────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def secretaire_view(request):
    produits = Produit.objects.all().order_by('nom')
    mouvements = Mouvement.objects.select_related('produit', 'utilisateur').order_by('-date')[:20]
    return render(request, 'gestion_stock/secretaire.html', {
        'produits': produits,
        'mouvements': mouvements,
    })


# ─── Utilisateurs ─────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def utilisateurs_view(request):
    if not request.user.is_superuser:
        messages.error(request, "Accès réservé à l'administrateur.")
        return redirect('home')
    utilisateurs = User.objects.exclude(is_superuser=True).prefetch_related('groups')
    groupes = Group.objects.all()
    return render(request, 'gestion_stock/utilisateurs.html', {
        'utilisateurs': utilisateurs,
        'groupes': groupes,
    })


@login_required(login_url='/login/')
def creer_utilisateur(request):
    if not request.user.is_superuser:
        return redirect('home')
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        groupe_id = request.POST.get('groupe')
        if not username or not password:
            messages.error(request, "Nom d'utilisateur et mot de passe requis.")
        elif User.objects.filter(username=username).exists():
            messages.error(request, f'L\'utilisateur "{username}" existe déjà.')
        else:
            user = User.objects.create_user(username=username, password=password)
            if groupe_id:
                try:
                    groupe = Group.objects.get(pk=groupe_id)
                    user.groups.add(groupe)
                except Group.DoesNotExist:
                    pass
            messages.success(request, f'Utilisateur "{username}" créé avec succès.')
    return redirect('utilisateurs')


@login_required(login_url='/login/')
def toggle_utilisateur(request, pk):
    if not request.user.is_superuser:
        return redirect('home')
    if request.method == 'POST':
        user = get_object_or_404(User, pk=pk)
        user.is_active = not user.is_active
        user.save()
        etat = 'activé' if user.is_active else 'désactivé'
        messages.success(request, f'Compte "{user.username}" {etat}.')
    return redirect('utilisateurs')


@login_required(login_url='/login/')
def supprimer_utilisateur(request, pk):
    if not request.user.is_superuser:
        return redirect('home')
    if request.method == 'POST':
        user = get_object_or_404(User, pk=pk)
        nom = user.username
        user.delete()
        messages.success(request, f'Utilisateur "{nom}" supprimé.')
    return redirect('utilisateurs')


# ─── Étape 4 : Exports ────────────────────────────────────────────────────────
import io
from datetime import datetime
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


@login_required(login_url='/login/')
def export_excel(request):
    produits = Produit.objects.all().order_by('nom')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventaire"

    # Styles
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(name='Calibri', bold=True, color="F0C040", size=11)
    accent_fill = PatternFill(start_color="F0C040", end_color="F0C040", fill_type="solid")
    accent_font = Font(name='Calibri', bold=True, color="1a1a2e", size=14)
    normal_font = Font(name='Calibri', size=10)
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    # Titre
    ws.merge_cells('A1:F1')
    ws['A1'] = 'RAPPORT D\'INVENTAIRE — STOCKPRO'
    ws['A1'].font = accent_font
    ws['A1'].fill = accent_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2')
    ws['A2'] = f'Généré le {datetime.now().strftime("%d/%m/%Y à %H:%M")}'
    ws['A2'].font = Font(name='Calibri', italic=True, color="888888", size=9)
    ws['A2'].alignment = Alignment(horizontal='center')

    # En-têtes
    headers = ['#', 'Produit', 'Quantité', 'Prix unitaire (FCFA)', 'Valeur totale (FCFA)', 'Statut']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[4].height = 22

    # Données
    valeur_totale = 0
    for i, p in enumerate(produits, 1):
        valeur = float(p.prix) * p.quantite
        valeur_totale += valeur
        if p.quantite == 0:
            statut = 'Épuisé'
        elif p.quantite < 5:
            statut = 'Stock critique'
        elif p.quantite < 10:
            statut = 'Stock faible'
        else:
            statut = 'OK'

        row = [i, p.nom, p.quantite, float(p.prix), valeur, statut]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=4 + i, column=col, value=val)
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col != 2 else 'left', vertical='center')
            if i % 2 == 0:
                cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")

    # Total
    total_row = 4 + len(list(produits)) + 2
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True, name='Calibri')
    ws.cell(row=total_row, column=5, value=valeur_totale).font = Font(bold=True, name='Calibri', color="1a6e3c")
    ws.cell(row=total_row, column=5).number_format = '#,##0.00'

    # Largeurs colonnes
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 15

    # Réponse HTTP
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"inventaire_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required(login_url='/login/')
def export_pdf(request):
    produits = Produit.objects.all().order_by('nom')
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    elements = []

    # Titre
    title_style = ParagraphStyle('title', parent=styles['Title'],
                                  fontSize=20, textColor=colors.HexColor('#1a1a2e'),
                                  spaceAfter=4)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                                fontSize=9, textColor=colors.grey, spaceAfter=20)

    elements.append(Paragraph("RAPPORT D'INVENTAIRE", title_style))
    elements.append(Paragraph(f"StockPro · Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", sub_style))

    # Tableau produits
    data = [['#', 'Produit', 'Quantité', 'Prix unitaire', 'Valeur totale', 'Statut']]
    valeur_totale = 0
    for i, p in enumerate(produits, 1):
        valeur = float(p.prix) * p.quantite
        valeur_totale += valeur
        if p.quantite == 0:
            statut = 'Épuisé'
        elif p.quantite < 5:
            statut = 'Critique'
        elif p.quantite < 10:
            statut = 'Faible'
        else:
            statut = 'OK'
        data.append([
            str(i), p.nom, str(p.quantite),
            f"{float(p.prix):,.0f} F",
            f"{valeur:,.0f} F",
            statut
        ])

    data.append(['', 'TOTAL', '', '', f"{valeur_totale:,.0f} F", ''])

    table = Table(data, colWidths=[1*cm, 6*cm, 2.5*cm, 3.5*cm, 3.5*cm, 2.5*cm])
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#F0C040')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ROWBACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        # Données
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('ROWBACKGROUND', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),
        # Total
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F0C040')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1a1a2e')),
        # Bordures
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('ROWHEIGHT', (0, 0), (-1, -1), 18),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    # Stats résumé
    elements.append(Spacer(1, 0.8*cm))
    total_produits = produits.count()
    alertes = produits.filter(quantite__lt=5).count()
    resume_style = ParagraphStyle('resume', parent=styles['Normal'], fontSize=9,
                                   textColor=colors.HexColor('#444444'), spaceAfter=4)
    elements.append(Paragraph(f"<b>Total produits :</b> {total_produits}", resume_style))
    elements.append(Paragraph(f"<b>Valeur totale du stock :</b> {valeur_totale:,.0f} FCFA", resume_style))
    elements.append(Paragraph(f"<b>Produits en stock critique (&lt;5) :</b> {alertes}", resume_style))

    doc.build(elements)
    buffer.seek(0)
    filename = f"inventaire_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
